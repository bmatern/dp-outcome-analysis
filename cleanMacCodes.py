import argparse
#from os import makedirs
#from os.path import join, isdir

from openpyxl import load_workbook
import pyard



def parseArgs():
    print('Parsing commandline arguments..')
    parser = argparse.ArgumentParser()

    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-x", "--excel", required=True, help="excel input file name. Row 1=Headers. Row2:X=HLA Typings", type=str)
    parser.add_argument("-c", "--columns", required=True, help="Columns that contain HLA typings with MAC Codes, separate by commas ex. G,H", type=str)
    parser.add_argument("-m", "--mode", required=False, help="Mode of operation to export all alleles or just the first one, options ('ALL','FIRST')", default='ALL', type=str)
    parser.add_argument("-t", "--twofield", help="Reduce all results to two field typing", action="store_true")

    args = parser.parse_args()

    return args

def isInteger(text=None):
    try:
        int(text)
        return True
    except ValueError:
        return False


def reduceFields(cellData=None):
    if not args.twofield:
        return cellData
    else:
        tokens = cellData.split(':')
        if len(tokens) > 2 :
            return tokens[0] + ':' + tokens[1]
        else:
            return cellData


def convertAlleleString(cellData=None, locus=None, ardObject=None):
    try:
        if(cellData is None or len(cellData) < 1):
            return ''
        #print('Converting cellData:' + str(cellData))

        nomenclatureTokens = cellData.split(':')
        alleleString = locus + '*' + cellData
        #print('allele string:' + str(alleleString))


        if (len(nomenclatureTokens)==2 and isInteger(nomenclatureTokens[0]) and nomenclatureTokens[1] == 'XX'):
            #print('This is an XX allele.')
            expandedXalleles = ardObject.redux_gl(alleleString, 'lgx').split('/')

            # Get rid of the locus again. Not sure if this is correct.
            for alleleIndex, allele in enumerate(expandedXalleles):
                expandedXalleles[alleleIndex] = allele.split('*')[1]

            return '|'.join(expandedXalleles)
        elif (len(nomenclatureTokens) == 2 and isInteger(nomenclatureTokens[0]) and not isInteger(nomenclatureTokens[1]) and nomenclatureTokens[1] != 'XX'):
            # is it a P group?
            # I suppose this could be repeated for G group.
            # This is only checking the second field to look for a P group, ex. "23:01P"
            if(alleleString.endswith('P') and isInteger(nomenclatureTokens[1][:-1])):
                print('P group:' + str(alleleString))
                # TODO: This actually doesnt work, the expansion of P groups does not get the entire list of possible alleles. This excludes some of the (rare) allele options.
                expandedPalleles = ardObject.redux_gl(alleleString, 'lgx').split('/')

                # Get rid of the locus again. Not sure if this is correct.
                for alleleIndex, allele in enumerate(expandedPalleles):
                    expandedPalleles[alleleIndex] = allele.split('*')[1]

                return '|'.join(expandedPalleles)



            #print('This is a MAC Code.')
            expandMac = ardObject.expand_mac(alleleString)

            # Get rid of the locus again. Not sure if this is correct.
            for alleleIndex, allele in enumerate(expandMac):
                expandMac[alleleIndex] = allele.split('*')[1]

            return '|'.join(expandMac)
        else:
            # This should be a normal allele. Re-return it.
            return reduceFields(cellData)

    except Exception as e:
        print('Exception:' + str(e))
        raise Exception('What should i do with this cellData:' + str(cellData))
        #expandMac = ardObject.expand_mac(alleleString)
        #print('Warning!!!!!!!!! this cell data had an issue, i will not handle it:' + str(cellData))


def loadExcelFile(excelFileName=None):
    if (verbose):
        print('Loading Donor File:' + str(excelFileName))
    workBook = load_workbook(excelFileName)
    return workBook


def exportFile(outputFileName=None, cleanedWorkbook=None):
    #outputFileName=join(outputDirectory,'ResultWithMatchScore.xlsx')
    if (verbose):
        print('Exporting Donor File:' + str(outputFileName))
    cleanedWorkbook.save(outputFileName)


def getArdObject():
    if (verbose):
        print('Creating ARD object for latest IMGT/HLA Release..')
    ard = pyard.ARD()
    return ard


def parseLocusName(rawLocusName=None, delimiter='-'):
    try:
        locusTokens = str(rawLocusName).strip().upper().split(delimiter)
        locus = locusTokens[0] + '-' + locusTokens[1]
        return locus
    except Exception as e:
        print('exception:' + str(e))
        print('rawLocusName:' + str(rawLocusName))
        raise e


def cleanMacCodes(excelFileName=None, columns=None, delimiter=',', headers=True):
    columnList = columns.split(delimiter)
    if (verbose):
        print('Cleaning Mac Codes in file ' + excelFileName + ' for ' + str(len(columnList)) + ' columns: ' + str(columnList))
    excelData = loadExcelFile(excelFileName=excelFileName)
    ardObject = getArdObject()

    firstSheet=excelData[excelData.sheetnames[0]]
    if(headers):
        startRowIndex=2
    else:
        startRowIndex=1
    #endRowIndex=18 # For testing.
    endRowIndex=None

    for columnName in columnList:
        rawLocusName=str(firstSheet[str(columnName) + str(1)].value)
        print('Column ' + str(columnName) + ' has locus name ' + str(rawLocusName))
        locusName = parseLocusName(rawLocusName=rawLocusName)
        #print('Locus:' + str(locusName))

        for rowIndexRaw, row in enumerate(firstSheet.iter_rows(min_row=startRowIndex, max_row=endRowIndex, values_only=True)):
            excelRowNumber = rowIndexRaw + startRowIndex
            #print('Raw Index ' + str(rowIndexRaw) + ', Excel Index ' + str(excelRowNumber) + ':' + str(row))

            columnName=columnName.strip()
            #print('Checking Cell ' + str(columnName) + str(excelRowNumber))
            cellData=firstSheet[str(columnName) + str(excelRowNumber)].value
            #print('Found Data:' + str(cellData))
            cleanedCellData = convertAlleleString(cellData=cellData, locus=locusName, ardObject=ardObject)

            if(str(args.mode).upper() == 'FIRST'):
                firstSheet[str(columnName) + str(excelRowNumber)] = cleanedCellData.split('|')[0]
            elif(str(args.mode).upper() == 'ALL'):
                firstSheet[str(columnName) + str(excelRowNumber)] = cleanedCellData
            else:
                raise Exception ('I expected a mode of FIRST or ALL, something went wrong')




    return excelData







if __name__ == '__main__':
    args=parseArgs()

    verbose = args.verbose
    if verbose:
        print("verbose mode active")


    cleanedFileName = args.excel.replace('.xlsx','.cleanedMacCodes.xlsx')

    cleanedWorkbook = cleanMacCodes(excelFileName=args.excel, columns=args.columns)
    exportFile(outputFileName=cleanedFileName, cleanedWorkbook=cleanedWorkbook)

    print('All done.')
