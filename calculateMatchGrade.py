import argparse
from os import makedirs
from os.path import join, isdir

from openpyxl import load_workbook


def readNMDPCodes(macFileName=None):
    if(verbose):
        print('Reading NMDP MAC codes from file ' + str(macFileName))
    nmdpCodeLookup={}
    nmdpCodeFile = open(macFileName, 'r')
    nmdpCodeLines = nmdpCodeFile.readlines()
    #print('I found this many lines: ' + str(len(nmdpCodeLines)))
    #print('The first line is this one: ' + str(nmdpCodeLines[3]))
    for lineIndex, dataLine in enumerate(nmdpCodeLines):
        # progress bar..
        if(lineIndex % 200000 == 0):
            print('nmdp code line ' + str(lineIndex) + ' / ' + str(len(nmdpCodeLines)))


        # first 3 lines are header info.
        if (lineIndex >= 3):
            lineTokens = dataLine.replace('\n','').split('\t')
            if (len(lineTokens) == 2):
                nmdpCode = lineTokens[0]
                alleleListString = lineTokens[1]
                alleleList = alleleListString.split('/')
                nmdpCodeLookup[nmdpCode]=alleleList
            elif(len(lineTokens) == 1 and dataLine == '\n'):
                pass
                #print('I guess this is the last empty line in the file, no big deal.')
            else:
                raise Exception ('I only expected 2 tokens in this line, investigate what this data means:' + str(lineTokens))

    nmdpCodeFile=None
    nmdpCodeLines=None

    return nmdpCodeLookup


def parseArgs():
    print('Parsing commandline arguments..')
    parser = argparse.ArgumentParser()

    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-d", "--donor", required=True, help="donor excel matched input file name", type=str)
    parser.add_argument("-o", "--output", required=True, help="output directory", type=str)
    parser.add_argument("-m", "--mac", required=True, help="mac codes text inputfile", type=str)

    args = parser.parse_args()

    return args


def loadDonorFile(donorFileName=None):
    if (verbose):
        print('Loading Donor File:' + str(donorFileName))
    workBook = load_workbook(donorFileName)
    return workBook


def exportFile(outputDirectory=None, workbookWithMatchScores=None):
    outputFileName=join(outputDirectory,'ResultWithMatchScore.xlsx')
    if (verbose):
        print('Exporting Donor File:' + str(outputFileName))
    if not isdir(outputDirectory):
        makedirs(outputDirectory)
    workbookWithMatchScores.save(outputFileName)


def isInteger(text=None):
    try:
        int(text)
        return True
    except ValueError:
        return False

def interpretAlleleString(alleleString=None, macLookup=None):
    try:
        alleleTokens = alleleString.split(':')
        alleleOptions = set()

        # If first two fields are numbers, this is unambiguous.
        if(isInteger(alleleTokens[0]) and isInteger(alleleTokens[1])):
            alleleOptions.add(str(alleleTokens[0]) + ':' + str(alleleTokens[1]))
        # If second field is XX then we have some ambiguity there. It can be anything.
        elif (isInteger(alleleTokens[0]) and alleleTokens[1]=='XX'):
            alleleOptions.add(str(alleleTokens[0]) + ':' + str(alleleTokens[1]))
        # integer:text is an NMDP code, or possibly it's an allele with expression marker.
        elif(isInteger(alleleTokens[0]) and not isInteger(alleleTokens[1])):

            # is it an expression marker?
            expressionMarkers = ['N', 'Q']
            if(isInteger(alleleTokens[1][0:len(alleleTokens[1])-1]) and alleleTokens[1][len(alleleTokens[1])-1:len(alleleTokens[1])] in expressionMarkers):
                #print('Probably expression marker: ' + alleleTokens[1])
                #Just a single option here. good.
                alleleOptions.add(str(alleleTokens[0]) + ':' + str(alleleTokens[1]))
            else:
                # Mac code.
                ambiguities = macLookup[alleleTokens[1]]
                for ambiguity in ambiguities:
                    # Mac codes are a bit confusing.
                    # Two options, either single, normal allele lists (these represent 2nd field protein name)
                    # Or full two-field allelenames, which represent 1st and 2nd-field allele names
                    if(':' not in ambiguity):
                        alleleOptions.add(str(alleleTokens[0]) + ':' + ambiguity)
                    else:
                        macTokens=ambiguity.split(':')
                        if(len(macTokens)==2):
                            # Hopefully the allele group matches...
                            if(alleleTokens[0] != macTokens[0]):
                                # This happens rarely. an allele from a mac string doesnt seem to match
                                print('Warning:This is weird, i would expect the allele to match the ambiguity string:')
                                print('AlleleString:' + str(alleleString))
                                print('MacAmbiguity:' + str(ambiguity))
                                print('From MAC list:' + str(ambiguities))
                                print('I will ignore the mismatched allele.')

                            # Use the allele names from the mac code.
                            alleleOptions.add(str(macTokens[0]) + ':' + str(macTokens[1]))

                        else:
                            raise Exception('There arent two mac tokens in this string:' + str(ambiguity))
        else:
            raise Exception ('What should i do with this allele text:' + str(alleleString))
        return alleleOptions

    except Exception:
        raise Exception('What should i do with this allele text:' + str(alleleString))


def interpretTypings(spreadsheetRow=None, dataColumns=None, macLookup=None):
    # This method interprets two excel cells. I return two sets of potential typings.
    # In some cases I need to use "XX" as a wildcard. This is if data is missing or not typed.
    # XX can match with any allele. I guess.
    #print('Row=' + str(spreadsheetRow))
    #print('DataColumns=' + str(dataColumns))
    if(len(dataColumns) != 2):
        raise Exception('Please send two typings at a time.')

    dataAllele1 = str(spreadsheetRow[dataColumns[0]]).strip().replace('None','').replace('NEW','')
    dataAllele2 = str(spreadsheetRow[dataColumns[1]]).strip().replace('None','').replace('NEW','')
    #print('Typing1:' + dataAllele1)
    #print('Typing2:' + dataAllele2)

    # If they're both blank, return two wildcards
    if(dataAllele1=='' and dataAllele1==''):
        wildCardSet = set()
        wildCardSet.add('XX:XX')
        return wildCardSet,wildCardSet
    # If the second one is blank, this means homozygous. Interpret the same allele twice
    elif(dataAllele2==''):
        dataAllele2 = dataAllele1
    # Otherwise, we have some data for both.
    else:
        pass

    allele1Set = interpretAlleleString(alleleString=dataAllele1, macLookup=macLookup)
    allele2Set = interpretAlleleString(alleleString=dataAllele2, macLookup=macLookup)

    #print('Returning:' + str(allele1Set) + '&' + str(allele2Set))
    return allele1Set, allele2Set


def isMatch(typing1=None, typing2=None):
    # Easy answer, if there is set overlap then match.
    if(len(typing1.intersection(typing2))>0):
        return True

    for allele1 in list(typing1):
        for allele2 in list(typing2):
            if(allele1==allele2):
                # Don't think this will ever happen, we already checked intersect
                return True
            else:
                allele1Tokens = allele1.split(':')
                allele2Tokens = allele2.split(':')
                # If either is ambig in first field?
                if(allele1Tokens[0]=='XX' or allele2Tokens[0]=='XX'):
                    return True

                # If group matches and one is ambiguous?
                elif(allele1Tokens[0]==allele2Tokens[0] and (allele1Tokens[1]=='XX' or allele2Tokens[1]=='XX')):
                    return True

    return False



def calculateIndividualScores(patientTyping1=None, patientTyping2=None, donorTyping1=None, donorTyping2=None):
    #print('Patient1:' + str(patientTyping1))
    #print('Patient2:' + str(patientTyping2))
    #print('Donor1:' + str(donorTyping1))
    #print('Donor2:' + str(donorTyping2))
    currentScore=0
    if(isMatch(patientTyping1,donorTyping1) or isMatch(patientTyping1, donorTyping2)):
        currentScore += 1
    if (isMatch(patientTyping2, donorTyping1) or isMatch(patientTyping2, donorTyping2)):
        currentScore += 1

    #print('Score=' + str(currentScore))
    return currentScore


def calculateMatchScoreFromTypings(patientTypings=None, donorTypings=None):
    aMatchScore = calculateIndividualScores(patientTyping1=patientTypings['A_1'], patientTyping2=patientTypings['A_2'], donorTyping1=donorTypings['A_1'], donorTyping2=donorTypings['A_2'])
    bMatchScore = calculateIndividualScores(patientTyping1=patientTypings['B_1'], patientTyping2=patientTypings['B_2'], donorTyping1=donorTypings['B_1'], donorTyping2=donorTypings['B_2'])
    cMatchScore = calculateIndividualScores(patientTyping1=patientTypings['C_1'], patientTyping2=patientTypings['C_2'], donorTyping1=donorTypings['C_1'], donorTyping2=donorTypings['C_2'])
    drb1MatchScore = calculateIndividualScores(patientTyping1=patientTypings['DRB1_1'], patientTyping2=patientTypings['DRB1_2'], donorTyping1=donorTypings['DRB1_1'], donorTyping2=donorTypings['DRB1_2'])
    dqb1MatchScore = calculateIndividualScores(patientTyping1=patientTypings['DQB1_1'], patientTyping2=patientTypings['DQB1_2'], donorTyping1=donorTypings['DQB1_1'], donorTyping2=donorTypings['DQB1_2'])
    totalScore = aMatchScore + bMatchScore + cMatchScore + drb1MatchScore + dqb1MatchScore
    return totalScore


def calculateMatchScoresPerLine(matchedTypingWorkbook=None, macLookup=None):
    if (verbose):
        print('Calculating match score per transplantation..')

    #print (matchedTypingWorkbook.sheetnames)
    firstSheet=matchedTypingWorkbook[matchedTypingWorkbook.sheetnames[0]]

    columnNames = []
    for column in firstSheet.iter_cols(min_row=1, max_row=1, values_only=True):
        #print(column)
        columnNames.append(column[0])
    #print('columnnames=' + str(columnNames))

    # Columns are 0-indexed. Rows are 1-indexed. Okay.
    scoreColumn = 'AL'
    firstSheet[scoreColumn+'1']='Match_Score'
    startRowIndex=2
    endRowIndex=None
    #startRowIndex=332 # For Testing
    #endRowIndex=332
    for rowIndexRaw, row in enumerate(firstSheet.iter_rows(min_row=startRowIndex, max_row=endRowIndex, values_only=True)):
        #print('RawIndex:' + str(rowIndexRaw))
        actualSpreadsheetRow = startRowIndex + rowIndexRaw
        #print('ActualIndex:' + str(actualSpreadsheetRow))

        patientTypings = {}
        patientTypings['A_1'], patientTypings['A_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[6,7], macLookup=macLookup)
        patientTypings['B_1'], patientTypings['B_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[8, 9], macLookup=macLookup)
        patientTypings['C_1'], patientTypings['C_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[10, 11], macLookup=macLookup)
        patientTypings['DRB1_1'], patientTypings['DRB1_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[12, 13], macLookup=macLookup)
        patientTypings['DQB1_1'], patientTypings['DQB1_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[14, 15], macLookup=macLookup)

        donorTypings = {}
        donorTypings['A_1'], donorTypings['A_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[25,26], macLookup=macLookup)
        donorTypings['B_1'], donorTypings['B_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[27, 28], macLookup=macLookup)
        donorTypings['C_1'], donorTypings['C_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[29, 30], macLookup=macLookup)
        donorTypings['DRB1_1'], donorTypings['DRB1_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[31, 32], macLookup=macLookup)
        donorTypings['DQB1_1'], donorTypings['DQB1_2'] = interpretTypings(spreadsheetRow=row, dataColumns=[33, 34], macLookup=macLookup)

        matchScore=calculateMatchScoreFromTypings(patientTypings=patientTypings, donorTypings=donorTypings)
        firstSheet[scoreColumn + str(actualSpreadsheetRow)] = matchScore




    # 1) OPen file load typings
    # 2) Handle homozygous alleles
    # 3) Do Mac Lookups (some of them are lists of typings)
    # 4) For each locus (a,b,c,dr,dq)
    # Coun # matches 0,1,2 (Match if ANY of the 2 field typings match)
    # 5) Total (probably 8-10, maybe 0)
    # 6) Export it again.
    return matchedTypingWorkbook


def calculateMatchGrade(donorFileName=None, macFileName=None, outputDirectory=None):
    if(verbose):
        print('Calculating match grade\nDonorFile:' + str(donorFileName) + '\nMacFileName:' + str(macFileName) + '\nOutputDirectory:' + str(outputDirectory))

    matchedTypingWorkbook= loadDonorFile(donorFileName=donorFileName)
    macLookup = readNMDPCodes(macFileName=macFileName)
    workbookWithMatchScores = calculateMatchScoresPerLine(matchedTypingWorkbook=matchedTypingWorkbook, macLookup=macLookup)
    exportFile(outputDirectory=outputDirectory, workbookWithMatchScores=workbookWithMatchScores)



if __name__ == '__main__':
    args=parseArgs()

    verbose = args.verbose
    if verbose:
        print("verbose mode active")


    calculateMatchGrade(donorFileName=args.donor, macFileName=args.mac, outputDirectory=args.output)

    print('All done.')
